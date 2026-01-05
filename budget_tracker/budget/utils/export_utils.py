"""
Utility functions for exporting data to CSV and PDF
"""
import csv
from io import BytesIO, StringIO
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill


def export_transactions_to_csv(transactions):
    """
    Export transactions to CSV format
    """
    output = StringIO()
    writer = csv.writer(output)
    
    # Write header
    writer.writerow([
        'Date', 'Category', 'Type', 'Amount', 'Currency', 
        'Description', 'Shared Account'
    ])
    
    # Write data
    for transaction in transactions:
        writer.writerow([
            transaction.date.strftime('%Y-%m-%d'),
            transaction.category.name,
            transaction.category.type,
            str(transaction.amount),
            transaction.currency.code if transaction.currency else 'N/A',
            transaction.description,
            transaction.shared_account.name if transaction.shared_account else 'Personal'
        ])
    
    return output.getvalue()


def export_transactions_to_excel(transactions):
    """
    Export transactions to Excel format
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Transactions"
    
    # Define styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    # Write header
    headers = ['Date', 'Category', 'Type', 'Amount', 'Currency', 'Description', 'Shared Account']
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
    
    # Write data
    for row_num, transaction in enumerate(transactions, 2):
        ws.cell(row=row_num, column=1).value = transaction.date.strftime('%Y-%m-%d')
        ws.cell(row=row_num, column=2).value = transaction.category.name
        ws.cell(row=row_num, column=3).value = transaction.category.type
        ws.cell(row=row_num, column=4).value = float(transaction.amount)
        ws.cell(row=row_num, column=5).value = transaction.currency.code if transaction.currency else 'N/A'
        ws.cell(row=row_num, column=6).value = transaction.description
        ws.cell(row=row_num, column=7).value = transaction.shared_account.name if transaction.shared_account else 'Personal'
    
    # Adjust column widths
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column].width = adjusted_width
    
    # Save to BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    return output.getvalue()


def export_transactions_to_pdf(transactions, user):
    """
    Export transactions to PDF format
    """
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter)
    elements = []
    
    # Styles
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=24,
        textColor=colors.HexColor('#2C3E50'),
        spaceAfter=30,
        alignment=1  # Center
    )
    
    # Title
    title = Paragraph(f"Transaction Report - {user.username}", title_style)
    elements.append(title)
    
    # Date range
    date_text = Paragraph(
        f"Generated on: {datetime.now().strftime('%Y-%m-%d %H:%M')}",
        styles['Normal']
    )
    elements.append(date_text)
    elements.append(Spacer(1, 0.3*inch))
    
    # Prepare table data
    data = [['Date', 'Category', 'Type', 'Amount', 'Description']]
    
    for transaction in transactions:
        data.append([
            transaction.date.strftime('%Y-%m-%d'),
            transaction.category.name,
            transaction.category.type,
            f"{transaction.amount} {transaction.currency.code if transaction.currency else ''}",
            transaction.description[:30] + '...' if len(transaction.description) > 30 else transaction.description
        ])
    
    # Create table
    table = Table(data, colWidths=[1.2*inch, 1.5*inch, 1*inch, 1.2*inch, 2.5*inch])
    
    # Style table
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4472C4')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 12),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -1), 9),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey]),
    ]))
    
    elements.append(table)
    
    # Build PDF
    doc.build(elements)
    
    buffer.seek(0)
    return buffer.getvalue()


def export_financial_summary_to_pdf(summary_data, user):
    """
    Export financial summary to PDF format
    """
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4)
    elements = []
    
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=24,
        textColor=colors.HexColor('#2C3E50'),
        spaceAfter=30,
        alignment=1
    )
    
    # Title
    title = Paragraph(f"Financial Summary - {user.username}", title_style)
    elements.append(title)
    elements.append(Spacer(1, 0.3*inch))
    
    # Summary statistics
    summary_table_data = [
        ['Metric', 'Amount'],
        ['Total Income', f"${summary_data.get('total_income', 0):.2f}"],
        ['Total Expenses', f"${summary_data.get('total_expenses', 0):.2f}"],
        ['Net Balance', f"${summary_data.get('net_balance', 0):.2f}"],
    ]
    
    summary_table = Table(summary_table_data, colWidths=[3*inch, 2*inch])
    summary_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4472C4')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 14),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -1), 12),
    ]))
    
    elements.append(summary_table)
    elements.append(Spacer(1, 0.5*inch))
    
    # Expenses by category
    if summary_data.get('expenses_by_category'):
        elements.append(Paragraph("Expenses by Category", styles['Heading2']))
        elements.append(Spacer(1, 0.2*inch))
        
        category_data = [['Category', 'Amount']]
        for item in summary_data['expenses_by_category']:
            category_data.append([
                item['category'],
                f"${item['amount']:.2f}"
            ])
        
        category_table = Table(category_data, colWidths=[3*inch, 2*inch])
        category_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4472C4')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey]),
        ]))
        
        elements.append(category_table)
    
    doc.build(elements)
    buffer.seek(0)
    return buffer.getvalue()
